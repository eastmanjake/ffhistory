#!/usr/bin/env python3
"""
ff_history.py — Yahoo Fantasy Football multi-season history scraper

Loops over all seasons in SEASONS config, fetches all data via Yahoo API,
writes JSON to ~/ffhistory/docs/data/YYYY/ and optionally Excel to ~/ffhistory/output/ff_YYYY.xlsx.

Usage:
  python ff_history.py                        # run all seasons (JSON only)
  python ff_history.py --export               # run all seasons + write Excel
  python ff_history.py --season 2011          # run one season only
  python ff_history.py --season 2011 --export # one season + Excel

Resume logic: if ~/ffhistory/docs/data/{year}/standings.json already exists,
that season is skipped. To re-scrape, delete that season's folder and re-run.
"""

import os
import re
import sys
import json
import time
from datetime import datetime, timezone
import pandas as pd
import openpyxl
from requests_oauthlib import OAuth2Session

# ── SEASON CONFIG ─────────────────────────────────────────────────────────────
SEASONS = [
    {"year": 2011, "game_id": "257", "league_id": "163099", "num_weeks": 16, "playoff_start": 14},
    {"year": 2012, "game_id": "273", "league_id": "176570", "num_weeks": 16, "playoff_start": 14},
    {"year": 2013, "game_id": "314", "league_id": "222964", "num_weeks": 16, "playoff_start": 14},
    {"year": 2014, "game_id": "331", "league_id": "285789", "num_weeks": 16, "playoff_start": 14},
    {"year": 2015, "game_id": "348", "league_id": "284573", "num_weeks": 16, "playoff_start": 14},
    {"year": 2016, "game_id": "359", "league_id": "371414", "num_weeks": 16, "playoff_start": 14},
    {"year": 2017, "game_id": "371", "league_id": "351465", "num_weeks": 16, "playoff_start": 14},
    {"year": 2018, "game_id": "380", "league_id": "208659", "num_weeks": 16, "playoff_start": 14},
    {"year": 2019, "game_id": "390", "league_id": "500600", "num_weeks": 16, "playoff_start": 14},
    {"year": 2020, "game_id": "399", "league_id": "362789", "num_weeks": 16, "playoff_start": 14},
    {"year": 2021, "game_id": "406", "league_id": "237408", "num_weeks": 17, "playoff_start": 15},
    {"year": 2022, "game_id": "414", "league_id": "247217", "num_weeks": 17, "playoff_start": 15},
    {"year": 2023, "game_id": "423", "league_id": "251046", "num_weeks": 17, "playoff_start": 15},
    {"year": 2024, "game_id": "449", "league_id": "215072", "num_weeks": 17, "playoff_start": 15},
    {"year": 2025, "game_id": "461", "league_id": "833763", "num_weeks": 17, "playoff_start": 15},
]

# ── OWNER OVERRIDES ───────────────────────────────────────────────────────────
# Full team_key -> real name. Add entries after verifying each season's owner mapping.
# Use "?" suffix for guesses, "Unknown" if not known.
# Yahoo nicknames are used as-is for unambiguous owners (Andrew, Brett, John, Michael).
OWNER_OVERRIDES = {
    # 2011
    "257.l.163099.t.2": "Andre?",    # Jennys New Lover (hidden, guess)
    "257.l.163099.t.3": "Sean",      # 16-0 (hidden)
    "257.l.163099.t.4": "Richie?",   # old dirty bastard (hidden, guess)
    "257.l.163099.t.5": "Jacob S.",  # postseason champs (Yahoo: Jake)
    "257.l.163099.t.6": "Jake",      # Goatse III (Yahoo: Jake)
    "257.l.163099.t.8": "Unknown",   # Taint Lickers (hidden)
    # 2012
    "273.l.176570.t.7":  "Steve",    # Raaaaaaaandy
    "273.l.176570.t.10": "Ken",      # Kenny kil
    "273.l.176570.t.8":  "Andre",    # Flint City Tropics
    # 2013
    "314.l.222964.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    "314.l.222964.t.7": "Ken",       # Kenny kil (hidden)
    # 2014
    "331.l.285789.t.4": "Saiku",     # KOU KILL EM (hidden)
    "331.l.285789.t.7": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2015
    "348.l.284573.t.7": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    "348.l.284573.t.9": "Ben",       # Ben's Nice Team (hidden)
    # 2016
    "359.l.371414.t.7": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2017
    "371.l.351465.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2018
    "380.l.208659.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2019
    "390.l.500600.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2020
    "399.l.362789.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2021
    "406.l.237408.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2022
    "414.l.247217.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2023
    "423.l.251046.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2024
    "449.l.215072.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
    # 2025
    "461.l.833763.t.6": "Steve",     # Raaaaaaaandy (Yahoo: Steven)
}

# ── CREDENTIALS & PATHS ───────────────────────────────────────────────────────
CONSUMER_KEY       = "dj0yJmk9Q0pnTU1pNE42ako5JmQ9WVdrOVdraFZSMnAwWVZnbWNHbzlNQT09JnM9Y29uc3VtZXJzZWNyZXQmc3Y9MCZ4PTdh"
CONSUMER_SECRET    = "6717467220c40d91faba43ca37bc2e1a2982b7a6"
TOKEN_FILE         = os.path.expanduser("~/ffhistory/token.json")
REDIRECT_URI       = "https://localhost"
AUTHORIZATION_BASE_URL = "https://api.login.yahoo.com/oauth2/request_auth"
TOKEN_URL          = "https://api.login.yahoo.com/oauth2/get_token"
BASE               = "https://fantasysports.yahooapis.com/fantasy/v2"
DATA_DIR           = os.path.expanduser("~/ffhistory/docs/data")
OUTPUT_DIR         = os.path.expanduser("~/ffhistory/output")

# ── CLI ARGS ──────────────────────────────────────────────────────────────────
EXPORT      = "--export" in sys.argv
TARGET_YEAR = None
if "--season" in sys.argv:
    idx = sys.argv.index("--season")
    TARGET_YEAR = int(sys.argv[idx + 1])

# ── AUTH ──────────────────────────────────────────────────────────────────────
def get_token():
    oauth = OAuth2Session(CONSUMER_KEY, redirect_uri=REDIRECT_URI)
    authorization_url, state = oauth.authorization_url(AUTHORIZATION_BASE_URL)
    print("\n--- AUTHORIZATION REQUIRED ---")
    print("Open this URL in your browser:")
    print(authorization_url)
    print("\nAfter approving, paste the verifier code from the redirect URL.")
    verifier = input("Enter verifier: ").strip()
    token = oauth.fetch_token(TOKEN_URL, code=verifier, client_secret=CONSUMER_SECRET)
    with open(TOKEN_FILE, "w") as f:
        json.dump(token, f)
    print("Token saved.")
    return token

def load_token():
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE) as f:
            return json.load(f)
    return None

def get_session():
    token = load_token()
    if not token:
        token = get_token()
    return OAuth2Session(CONSUMER_KEY, token=token)

def yahoo_get(session, url, _retries=5):
    from oauthlib.oauth2.rfc6749.errors import TokenExpiredError
    for attempt in range(_retries):
        try:
            response = session.get(url, headers={"Accept": "application/json"})
        except TokenExpiredError:
            print("Token expired, re-authenticating...")
            token = get_token()
            session.__dict__.update(OAuth2Session(CONSUMER_KEY, token=token).__dict__)
            response = session.get(url, headers={"Accept": "application/json"})
        if response.status_code == 401:
            print("Token expired (401), re-authenticating...")
            token = get_token()
            session.__dict__.update(OAuth2Session(CONSUMER_KEY, token=token).__dict__)
            response = session.get(url, headers={"Accept": "application/json"})
        if response.status_code == 429 or not response.text.strip():
            wait = 30 * (attempt + 1)
            print(f"  Rate limited (empty response), waiting {wait}s...")
            time.sleep(wait)
            continue
        try:
            return response.json()
        except Exception:
            wait = 30 * (attempt + 1)
            print(f"  Bad response (status {response.status_code}), waiting {wait}s...")
            time.sleep(wait)
    raise RuntimeError(f"Yahoo API failed after {_retries} retries: {url}")

def fmt_ts(ts):
    """Convert Unix timestamp to YYYY-MM-DD string."""
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d")
    except:
        return str(ts)

def normalize_name(name):
    """Normalize player name for matching against nfl_data_py."""
    name = name.lower().strip()
    name = name.replace('.', '')
    name = re.sub(r'\s+(jr|sr|ii|iii|iv)$', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name

def build_nfl_team_lookup(year):
    """
    Load nfl_data_py weekly rosters for the given year.
    Returns (by_yahoo_id, by_name):
      by_yahoo_id: {yahoo_id_int: {week_int: team_str}}
      by_name:     {normalized_name: {week_int: team_str}}
    """
    try:
        import nfl_data_py as nfl
        rosters = nfl.import_weekly_rosters([year])
        by_yahoo_id, by_name = {}, {}
        for _, row in rosters.iterrows():
            week_val = row.get('week')
            team_val = row.get('team')
            if pd.isna(week_val) or pd.isna(team_val):
                continue
            week = int(week_val)
            team = str(team_val)
            yid = row.get('yahoo_id')
            try:
                if yid is not None and str(yid).strip() != '' and not pd.isna(yid):
                    by_yahoo_id.setdefault(int(yid), {})[week] = team
            except (ValueError, TypeError):
                pass
            pname = str(row.get('player_name') or '')
            norm = normalize_name(pname)
            if norm:
                by_name.setdefault(norm, {})[week] = team
        return by_yahoo_id, by_name
    except Exception as e:
        print(f"  Warning: nfl_data_py roster lookup unavailable for {year}: {e}")
        return {}, {}

def lookup_nfl_team(player_key, player_name, week, by_yahoo_id, by_name, fallback=""):
    """Look up a player's NFL team for a given week using nfl_data_py data."""
    if player_key:
        try:
            yid = int(player_key.split('.')[-1])
            if yid in by_yahoo_id:
                week_map = by_yahoo_id[yid]
                if week in week_map:
                    return week_map[week]
                nearest = min(week_map, key=lambda w: abs(w - week))
                return week_map[nearest]
        except (ValueError, IndexError):
            pass
    norm = normalize_name(player_name)
    if norm in by_name:
        week_map = by_name[norm]
        if week in week_map:
            return week_map[week]
        nearest = min(week_map, key=lambda w: abs(w - week))
        return week_map[nearest]
    return fallback

# ── SEASON SCRAPER ────────────────────────────────────────────────────────────
def scrape_season(session, cfg, export=False):
    year          = cfg["year"]
    game_id       = cfg["game_id"]
    league_id     = cfg["league_id"]
    num_weeks     = cfg["num_weeks"]
    playoff_start = cfg["playoff_start"]
    league_key    = f"{game_id}.l.{league_id}"

    season_data_dir = os.path.join(DATA_DIR, str(year))
    season_xlsx     = os.path.join(OUTPUT_DIR, f"ff_{year}.xlsx")

    os.makedirs(season_data_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"  SEASON {year}  ({league_key})")
    print(f"{'='*60}")

    # ── SETTINGS ──────────────────────────────────────────────────────────────
    print("Fetching league settings...")
    sdata     = yahoo_get(session, f"{BASE}/league/{league_key}/settings?format=json")
    _settings = sdata["fantasy_content"]["league"][1]["settings"][0]
    stat_names    = {}
    scoring_rules = {}
    for _s in _settings["stat_categories"]["stats"]:
        stat_names[int(_s["stat"]["stat_id"])] = _s["stat"]["name"]
    for _s in _settings["stat_modifiers"]["stats"]:
        scoring_rules[int(_s["stat"]["stat_id"])] = float(_s["stat"]["value"])
    settings_rows = [{"stat": stat_names.get(sid, f"Stat {sid}"), "points": val}
                     for sid, val in scoring_rules.items()]

    # ── 1. STANDINGS ──────────────────────────────────────────────────────────
    print("Fetching standings...")
    data      = yahoo_get(session, f"{BASE}/league/{league_key}/standings?format=json")
    teams_raw = data["fantasy_content"]["league"][1]["standings"][0]["teams"]
    team_count = teams_raw["count"]

    team_key_to_name    = {}
    team_key_to_manager = {}
    standings_rows      = []

    for i in range(team_count):
        t        = teams_raw[str(i)]["team"]
        info     = t[0]
        stg      = t[2]["team_standings"]
        name     = next((x["name"] for x in info if isinstance(x, dict) and "name" in x), "")
        team_key = next((x["team_key"] for x in info if isinstance(x, dict) and "team_key" in x), "")
        manager  = ""
        for x in info:
            if isinstance(x, dict) and "managers" in x:
                manager = x["managers"][0]["manager"].get("nickname", "")
        owner = OWNER_OVERRIDES.get(team_key) or manager or "Unknown"
        team_key_to_name[team_key]    = name
        team_key_to_manager[team_key] = owner
        outcomes = stg["outcome_totals"]
        standings_rows.append({
            "team": name, "owner": owner,
            "wins": outcomes["wins"], "losses": outcomes["losses"], "ties": outcomes["ties"],
            "points_for": stg["points_for"], "points_against": stg["points_against"],
            "rank": stg["rank"]
        })

    print(f"\n  Owner mapping for {year}:")
    print(f"  {'Team Key':<25} {'Team Name':<28} Owner")
    print("  " + "-" * 70)
    for tk, owner in team_key_to_manager.items():
        print(f"  {tk:<25} {team_key_to_name.get(tk, tk):<28} {owner}")

    print(f"\n  Standings:")
    print(f"  {'Rank':>4}  {'Team':<28} {'Owner':<18} {'W':>3} {'L':>3} {'PF':>8} {'PA':>8}")
    print("  " + "-" * 76)
    for r in sorted(standings_rows, key=lambda x: int(x["rank"])):
        print(f"  {r['rank']:>4}  {r['team']:<28} {r['owner']:<18} {r['wins']:>3} "
              f"{r['losses']:>3} {float(r['points_for']):>8.2f} {float(r['points_against']):>8.2f}")

    # ── 2. DRAFT ──────────────────────────────────────────────────────────────
    print("\nFetching draft results...")
    data       = yahoo_get(session, f"{BASE}/league/{league_key}/draftresults?format=json")
    picks_raw  = data["fantasy_content"]["league"][1]["draft_results"]
    pick_count = picks_raw["count"]

    raw_picks   = []
    player_keys = []
    for i in range(pick_count):
        pick = picks_raw[str(i)]["draft_result"]
        raw_picks.append(pick)
        player_keys.append(pick["player_key"])

    print(f"  Resolving {len(player_keys)} player names...")
    player_key_to_info = {}
    for i in range(0, len(player_keys), 25):
        batch    = player_keys[i:i+25]
        keys_str = ",".join(batch)
        pdata    = yahoo_get(session, f"{BASE}/players;player_keys={keys_str}?format=json")
        players  = pdata["fantasy_content"]["players"]
        for j in range(players["count"]):
            p    = players[str(j)]["player"][0]
            pkey = next((x["player_key"] for x in p if isinstance(x, dict) and "player_key" in x), "")
            pname, ppos = "", ""
            for x in p:
                if isinstance(x, dict) and "name" in x and isinstance(x["name"], dict):
                    pname = x["name"].get("full", "")
                if isinstance(x, dict) and "display_position" in x:
                    ppos = x["display_position"]
            player_key_to_info[pkey] = (pname, ppos)
        print(f"  Fetched players {i+1}–{min(i+25, len(player_keys))}")

    draft_rows = []
    for pick in raw_picks:
        pname, ppos = player_key_to_info.get(pick["player_key"], (pick["player_key"], ""))
        draft_rows.append({
            "round": pick["round"], "pick": pick["pick"],
            "team": team_key_to_name.get(pick["team_key"], pick["team_key"]),
            "player": pname, "position": ppos
        })

    # ── 3. MATCHUPS ───────────────────────────────────────────────────────────
    print("\nFetching weekly matchups...")
    matchup_rows = []
    for week in range(1, num_weeks + 1):
        print(f"  Week {week}...")
        try:
            data          = yahoo_get(session, f"{BASE}/league/{league_key}/scoreboard;week={week}?format=json")
            matchups      = data["fantasy_content"]["league"][1]["scoreboard"]["0"]["matchups"]
            matchup_count = matchups["count"]
            for i in range(matchup_count):
                m       = matchups[str(i)]["matchup"]
                teams   = m["0"]["teams"]
                t1      = teams["0"]["team"]
                t2      = teams["1"]["team"]
                t1_name = next((x["name"] for x in t1[0] if isinstance(x, dict) and "name" in x), "")
                t2_name = next((x["name"] for x in t2[0] if isinstance(x, dict) and "name" in x), "")
                matchup_rows.append({
                    "week": week,
                    "team1": t1_name, "score1": t1[1]["team_points"]["total"],
                    "team2": t2_name, "score2": t2[1]["team_points"]["total"],
                    "playoff": week >= playoff_start,
                    "round": ""
                })
        except Exception as e:
            print(f"  Week {week} error: {e}")

    # Playoff bracket labeling
    seeded_teams  = [r["team"] for r in sorted(standings_rows,
                     key=lambda r: (-int(r["wins"]), -float(r["points_for"])))]
    playoff_teams = set(seeded_teams[:6])
    top_two       = set(seeded_teams[:2])
    week_winners  = {}
    week_losers   = {}

    for row in matchup_rows:
        if not row["playoff"]:
            continue
        week = row["week"]
        try:
            s1, s2 = float(row["score1"]), float(row["score2"])
        except:
            continue
        t1, t2 = row["team1"], row["team2"]
        winner = t1 if s1 > s2 else t2
        loser  = t1 if s1 < s2 else t2

        if week == playoff_start:
            if t1 in top_two or t2 in top_two:
                label = "Consolation"
            elif t1 in playoff_teams and t2 in playoff_teams:
                label = "Quarterfinal"
                week_winners.setdefault(week, set()).add(winner)
                week_losers.setdefault(week, set()).add(loser)
            else:
                label = "Consolation"
        elif week == playoff_start + 1:
            sf_teams = top_two | week_winners.get(playoff_start, set())
            if t1 in sf_teams and t2 in sf_teams:
                label = "Semifinal"
                week_winners.setdefault(week, set()).add(winner)
                week_losers.setdefault(week, set()).add(loser)
            else:
                label = "Consolation"
        elif week == playoff_start + 2:
            sf_winners = week_winners.get(playoff_start + 1, set())
            sf_losers  = week_losers.get(playoff_start + 1, set())
            if t1 in sf_winners and t2 in sf_winners:
                label = "Championship"
            elif t1 in sf_losers and t2 in sf_losers:
                label = "3rd Place"
            else:
                label = "Consolation"
        else:
            label = ""
        row["round"] = label

    print("\n  Playoff matchups:")
    print(f"  {'Wk':>3}  {'Team 1':<28} {'Score 1':>8}  {'Team 2':<28} {'Score 2':>8}  Round")
    print("  " + "-" * 92)
    for r in [r for r in matchup_rows if r["playoff"]]:
        print(f"  {r['week']:>3}  {r['team1']:<28} {float(r['score1']):>8.2f}  "
              f"{r['team2']:<28} {float(r['score2']):>8.2f}  {r['round']}")

    # ── 4. TRANSACTIONS ───────────────────────────────────────────────────────
    print("\nFetching transactions...")
    transaction_rows = []
    for t_type in ["add", "drop", "trade"]:
        print(f"  Fetching {t_type}s...")
        try:
            data         = yahoo_get(session, f"{BASE}/league/{league_key}/transactions;type={t_type}?format=json")
            transactions = data["fantasy_content"]["league"][1]["transactions"]
            if isinstance(transactions, list):
                transactions = transactions[0] if transactions else {}
            t_count      = transactions.get("count", 0) if isinstance(transactions, dict) else 0
            print(f"  Found {t_count} {t_type}s")
            for i in range(t_count):
                t         = transactions[str(i)]["transaction"]
                t_info    = t[0]
                t1_raw    = t[1] if len(t) > 1 else {}
                t_players = (t1_raw if isinstance(t1_raw, dict) else {}).get("players", {})
                timestamp = fmt_ts(t_info.get("timestamp", ""))
                for j in range(t_players.get("count", 0)):
                    p      = t_players[str(j)]["player"]
                    p_info = p[0]
                    p_data = p[1].get("transaction_data", {})
                    pname  = ""
                    for x in p_info:
                        if isinstance(x, dict) and "name" in x and isinstance(x["name"], dict):
                            pname = x["name"].get("full", "")
                    if isinstance(p_data, list):
                        p_data = p_data[0] if p_data else {}
                    transaction_rows.append({
                        "type": t_type.capitalize(),
                        "date": timestamp,
                        "player": pname,
                        "from": p_data.get("source_team_name", "FA/Waiver"),
                        "to":   p_data.get("destination_team_name", "Dropped"),
                    })
            time.sleep(0.5)
        except Exception as e:
            print(f"  Error fetching {t_type}s: {e}")

    # ── NFL TEAM LOOKUP ───────────────────────────────────────────────────────
    nfl_by_yahoo_id, nfl_by_name = build_nfl_team_lookup(year)

    # ── 5. END OF SEASON ROSTERS ──────────────────────────────────────────────
    print("\nFetching end of season rosters...")
    roster_rows = []
    for team_key, team_name in team_key_to_name.items():
        try:
            data    = yahoo_get(session, f"{BASE}/team/{team_key}/roster;week={num_weeks}?format=json")
            players = data["fantasy_content"]["team"][1]["roster"]["0"]["players"]
            for i in range(players["count"]):
                p = players[str(i)]["player"][0]
                pk, pname, ppos, pteam, pstatus = "", "", "", "", ""
                for x in p:
                    if isinstance(x, dict) and "player_key" in x:
                        pk = x["player_key"]
                    if isinstance(x, dict) and "name" in x and isinstance(x["name"], dict):
                        pname = x["name"].get("full", "")
                    if isinstance(x, dict) and "display_position" in x:
                        ppos = x["display_position"]
                    if isinstance(x, dict) and "editorial_team_abbr" in x:
                        pteam = x["editorial_team_abbr"]
                    if isinstance(x, dict) and "status" in x:
                        pstatus = x["status"]
                nfl_team = lookup_nfl_team(pk, pname, num_weeks, nfl_by_yahoo_id, nfl_by_name, pteam)
                roster_rows.append({
                    "season": year, "team": team_name,
                    "owner": team_key_to_manager.get(team_key, "Unknown"),
                    "player": pname, "position": ppos,
                    "nfl_team": nfl_team, "status": pstatus
                })
            print(f"  {team_name}: {players['count']} players")
            time.sleep(0.5)
        except Exception as e:
            print(f"  Error fetching roster for {team_name}: {e}")

    # ── 6. WEEKLY LINEUPS + FANTASY POINTS ────────────────────────────────────
    print("\nFetching weekly lineups + fantasy points...")
    lineup_rows = []
    for week in range(1, num_weeks + 1):
        print(f"  Week {week}...")
        for team_key, team_name in team_key_to_name.items():
            try:
                roster_data    = yahoo_get(session, f"{BASE}/team/{team_key}/roster;week={week}?format=json")
                roster_players = roster_data["fantasy_content"]["team"][1]["roster"]["0"]["players"]
                players_this_week = []
                for i in range(roster_players["count"]):
                    p          = roster_players[str(i)]["player"]
                    info       = p[0]
                    p_selected = p[1].get("selected_position", [])
                    pk    = next((x["player_key"] for x in info if isinstance(x, dict) and "player_key" in x), None)
                    pname = next((x["name"]["full"] for x in info if isinstance(x, dict) and "name" in x and isinstance(x["name"], dict)), "")
                    ppos  = next((x["display_position"] for x in info if isinstance(x, dict) and "display_position" in x), "")
                    pteam = next((x["editorial_team_abbr"] for x in info if isinstance(x, dict) and "editorial_team_abbr" in x), "")
                    slot  = ""
                    if isinstance(p_selected, list):
                        for s in p_selected:
                            if isinstance(s, dict) and "position" in s:
                                slot = s["position"]
                    elif isinstance(p_selected, dict):
                        slot = p_selected.get("position", "")
                    if pk:
                        players_this_week.append((pk, slot, pname, ppos, pteam))
                time.sleep(0.2)

                keys_str    = ",".join(pk for pk, *_ in players_this_week)
                pts_data    = yahoo_get(session, f"{BASE}/league/{league_key}/players;player_keys={keys_str}/stats;type=week;week={week}?format=json")
                pts_players = pts_data["fantasy_content"]["league"][1]["players"]
                pts_by_key  = {}
                for i in range(pts_players["count"]):
                    p    = pts_players[str(i)]["player"]
                    info = p[0]
                    sb   = p[1] if len(p) > 1 else {}
                    pk   = next((x["player_key"] for x in info if isinstance(x, dict) and "player_key" in x), "")
                    pts_by_key[pk] = sb.get("player_points", {}).get("total", "")
                time.sleep(0.2)

                owner = team_key_to_manager.get(team_key, "Unknown")
                for pk, slot, pname, ppos, pteam in players_this_week:
                    nfl_team = lookup_nfl_team(pk, pname, week, nfl_by_yahoo_id, nfl_by_name, pteam)
                    lineup_rows.append({
                        "week": week, "team": team_name, "owner": owner,
                        "slot": slot, "player": pname, "pos": ppos,
                        "nfl_team": nfl_team, "fantasy_pts": pts_by_key.get(pk, "")
                    })
            except Exception as e:
                print(f"    Error {team_name} week {week}: {e}")

    print(f"  {len(lineup_rows)} player-week records fetched.")

    # ── WRITE JSON ────────────────────────────────────────────────────────────
    print("\nWriting JSON...")
    json_files = {
        "standings.json":      standings_rows,
        "draft.json":          draft_rows,
        "matchups.json":       matchup_rows,
        "transactions.json":   transaction_rows,
        "rosters.json":        roster_rows,
        "weekly_lineups.json": lineup_rows,
        "settings.json":       settings_rows,
    }
    for fname, fdata in json_files.items():
        path = os.path.join(season_data_dir, fname)
        with open(path, "w") as f:
            json.dump(fdata, f, indent=2)
        print(f"  Wrote {path}")

    # ── WRITE EXCEL ───────────────────────────────────────────────────────────
    if not export:
        print("  (run with --export to also write Excel)")
        return

    print("Writing Excel...")
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Standings"
    ws.append(["Team", "Owner", "Wins", "Losses", "Ties", "Points For", "Points Against", "Rank"])
    for r in standings_rows:
        ws.append([r["team"], r["owner"], r["wins"], r["losses"], r["ties"],
                   r["points_for"], r["points_against"], r["rank"]])

    ws2 = wb.create_sheet("Draft")
    ws2.append(["Round", "Pick", "Team", "Player", "Position"])
    for r in draft_rows:
        ws2.append([r["round"], r["pick"], r["team"], r["player"], r["position"]])

    ws3 = wb.create_sheet("Matchups")
    ws3.append(["Week", "Team 1", "Score 1", "Team 2", "Score 2", "Playoff", "Round"])
    for r in matchup_rows:
        ws3.append([r["week"], r["team1"], r["score1"], r["team2"], r["score2"],
                    "Yes" if r["playoff"] else "No", r["round"]])

    ws4 = wb.create_sheet("Transactions")
    ws4.append(["Type", "Date", "Player", "From", "To"])
    for r in transaction_rows:
        ws4.append([r["type"], r["date"], r["player"], r["from"], r["to"]])

    ws5 = wb.create_sheet("End of Season Rosters")
    ws5.append(["Season", "Fantasy Team", "Owner", "Player", "Position", "NFL Team", "Status"])
    for r in roster_rows:
        ws5.append([r["season"], r["team"], r["owner"], r["player"],
                    r["position"], r["nfl_team"], r["status"]])

    ws6 = wb.create_sheet("Playoffs")
    ws6.append(["Week", "Team 1", "Score 1", "Team 2", "Score 2", "Round"])
    for r in matchup_rows:
        if r["playoff"]:
            ws6.append([r["week"], r["team1"], r["score1"], r["team2"], r["score2"], r["round"]])

    ws7 = wb.create_sheet("League Settings")
    ws7.append(["Stat", "Points"])
    for r in settings_rows:
        ws7.append([r["stat"], r["points"]])

    for week in range(1, num_weeks + 1):
        ws_week = wb.create_sheet(f"Week {week:02d}")
        ws_week.append(["Week", "Fantasy Team", "Owner", "Slot", "Player", "Pos", "NFL Team", "Fantasy Pts"])
        for r in [x for x in lineup_rows if x["week"] == week]:
            ws_week.append([r["week"], r["team"], r["owner"], r["slot"],
                            r["player"], r["pos"], r["nfl_team"], r["fantasy_pts"]])

    wb.save(season_xlsx)
    print(f"  Saved to {season_xlsx}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    session = get_session()

    seasons_to_run = [s for s in SEASONS if TARGET_YEAR is None or s["year"] == TARGET_YEAR]

    completed = []
    skipped   = []
    failed    = []

    for cfg in seasons_to_run:
        year            = cfg["year"]
        resume_sentinel = os.path.join(DATA_DIR, str(year), "standings.json")

        if os.path.exists(resume_sentinel):
            print(f"\nSkipping {year} — already complete (delete {DATA_DIR}/{year}/ to re-scrape)")
            skipped.append(year)
            continue

        try:
            scrape_season(session, cfg, export=EXPORT)
            completed.append(year)
        except Exception as e:
            print(f"\nERROR scraping {year}: {e}")
            failed.append(year)

    print("\n" + "="*60)
    print("  RUN SUMMARY")
    print("="*60)
    if completed:
        print(f"  Completed : {', '.join(str(y) for y in completed)}")
    if skipped:
        print(f"  Skipped   : {', '.join(str(y) for y in skipped)}")
    if failed:
        print(f"  Failed    : {', '.join(str(y) for y in failed)}")
    if not completed and not failed:
        print("  Nothing new to scrape.")
    print("="*60)


if __name__ == "__main__":
    main()
