import os
import sys
import json
import time
import openpyxl
from requests_oauthlib import OAuth2Session

EXPORT = "--export" in sys.argv

# ── LOAD LEAGUE SETTINGS + BUILD SCORING RULES ───────────────────────────────
SETTINGS_FILE = os.path.join(os.path.dirname(__file__), "settings_raw.json")
with open(SETTINGS_FILE) as _f:
    _raw = json.load(_f)
_settings = _raw["fantasy_content"]["league"][1]["settings"][0]

# stat_id -> display name and point value (for the settings tab)
stat_names = {}
for _s in _settings["stat_categories"]["stats"]:
    stat_names[int(_s["stat"]["stat_id"])] = _s["stat"]["name"]

scoring_rules = {}
for _s in _settings["stat_modifiers"]["stats"]:
    scoring_rules[int(_s["stat"]["stat_id"])] = float(_s["stat"]["value"])

# Rows for the League Settings tab: (Stat, Points)
settings_rows = [[stat_names.get(sid, f"Stat {sid}"), val] for sid, val in scoring_rules.items()]

# Your credentials
CONSUMER_KEY = "dj0yJmk9Q0pnTU1pNE42ako5JmQ9WVdrOVdraFZSMnAwWVZnbWNHbzlNQT09JnM9Y29uc3VtZXJzZWNyZXQmc3Y9MCZ4PTdh"
CONSUMER_SECRET = "6717467220c40d91faba43ca37bc2e1a2982b7a6"
LEAGUE_ID = "163099"
GAME_ID = "257"
SEASON_YEAR = 2011
PLAYOFF_START_WEEK = 14
NUM_WEEKS = 16

TOKEN_FILE = os.path.expanduser("~/ffhistory/token.json")
REDIRECT_URI = "https://localhost"
AUTHORIZATION_BASE_URL = "https://api.login.yahoo.com/oauth2/request_auth"
TOKEN_URL = "https://api.login.yahoo.com/oauth2/get_token"
BASE = "https://fantasysports.yahooapis.com/fantasy/v2"
LEAGUE_KEY = f"{GAME_ID}.l.{LEAGUE_ID}"

def get_token():
    oauth = OAuth2Session(CONSUMER_KEY, redirect_uri=REDIRECT_URI)
    authorization_url, state = oauth.authorization_url(AUTHORIZATION_BASE_URL)
    print("\n--- AUTHORIZATION REQUIRED ---")
    print("Open this URL in your browser:")
    print(authorization_url)
    print("\nAfter approving, you'll see a verification code.")
    verifier = input("Enter verifier: ").strip()
    token = oauth.fetch_token(TOKEN_URL, code=verifier, client_secret=CONSUMER_SECRET)
    with open(TOKEN_FILE, "w") as f:
        json.dump(token, f)
    print("Token saved!")
    return token

def load_token():
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE) as f:
            return json.load(f)
    return None

def get_session():
    token = load_token()
    if not token:
        token = get_token()
    return OAuth2Session(CONSUMER_KEY, token=token)

def yahoo_get(session, url):
    response = session.get(url, headers={"Accept": "application/json"})
    if response.status_code == 401:
        print("Token expired, re-authenticating...")
        token = get_token()
        session = OAuth2Session(CONSUMER_KEY, token=token)
        response = session.get(url, headers={"Accept": "application/json"})
    return response.json()

session = get_session()

# ── 1. STANDINGS ──────────────────────────────────────────────────────────────
print("Fetching standings...")
data = yahoo_get(session, f"{BASE}/league/{LEAGUE_KEY}/standings?format=json")
teams_raw = data["fantasy_content"]["league"][1]["standings"][0]["teams"]
team_count = teams_raw["count"]

standings_rows = []
for i in range(team_count):
    t = teams_raw[str(i)]["team"]
    info = t[0]
    standings = t[2]["team_standings"]
    name = next((x["name"] for x in info if isinstance(x, dict) and "name" in x), "")
    manager = ""
    for x in info:
        if isinstance(x, dict) and "managers" in x:
            manager = x["managers"][0]["manager"].get("nickname", "")
    outcomes = standings["outcome_totals"]
    standings_rows.append([
        name, manager,
        outcomes["wins"], outcomes["losses"], outcomes["ties"],
        standings["points_for"], standings["points_against"],
        standings["rank"]
    ])

print("\n=== STANDINGS ===")
print(f"{'Rank':>4}  {'Team':<28} {'Manager':<18} {'W':>3} {'L':>3} {'T':>3} {'PF':>8} {'PA':>8}")
print("-" * 80)
for r in sorted(standings_rows, key=lambda x: int(x[7])):
    print(f"{r[7]:>4}  {r[0]:<28} {r[1]:<18} {r[2]:>3} {r[3]:>3} {r[4]:>3} {float(r[5]):>8.2f} {float(r[6]):>8.2f}")

# ── 2. DRAFT RESULTS ─────────────────────────────────────────────────────────
print("\nFetching draft results...")
data = yahoo_get(session, f"{BASE}/league/{LEAGUE_KEY}/draftresults?format=json")
picks_raw = data["fantasy_content"]["league"][1]["draft_results"]
pick_count = picks_raw["count"]

# Build team key -> team name map from standings data
team_key_to_name = {}
for i in range(team_count):
    t = teams_raw[str(i)]["team"]
    info = t[0]
    name = next((x["name"] for x in info if isinstance(x, dict) and "name" in x), "")
    team_key = next((x["team_key"] for x in info if isinstance(x, dict) and "team_key" in x), "")
    team_key_to_name[team_key] = name

# Collect all pick data and player keys
raw_picks = []
player_keys = []
for i in range(pick_count):
    pick = picks_raw[str(i)]["draft_result"]
    raw_picks.append(pick)
    player_keys.append(pick["player_key"])

# Batch fetch player names (25 at a time)
print(f"  Resolving {len(player_keys)} player names...")
player_key_to_name = {}
batch_size = 25
for i in range(0, len(player_keys), batch_size):
    batch = player_keys[i:i+batch_size]
    keys_str = ",".join(batch)
    url = f"{BASE}/players;player_keys={keys_str}?format=json"
    pdata = yahoo_get(session, url)
    players = pdata["fantasy_content"]["players"]
    pcount = players["count"]
    for j in range(pcount):
        p = players[str(j)]["player"][0]
        pkey = next((x["player_key"] for x in p if isinstance(x, dict) and "player_key" in x), "")
        pname = ""
        pposition = ""
        for x in p:
            if isinstance(x, dict) and "name" in x and isinstance(x["name"], dict):
                pname = x["name"].get("full", "")
            if isinstance(x, dict) and "display_position" in x:
                pposition = x["display_position"]
        player_key_to_name[pkey] = (pname, pposition)
    print(f"  Fetched players {i+1} to {min(i+batch_size, len(player_keys))}")

# Build final draft rows
draft_rows = []
for pick in raw_picks:
    team_name = team_key_to_name.get(pick["team_key"], pick["team_key"])
    player_info = player_key_to_name.get(pick["player_key"], (pick["player_key"], ""))
    player_name, player_position = player_info if isinstance(player_info, tuple) else (player_info, "")
    draft_rows.append([
        pick["round"],
        pick["pick"],
        team_name,
        player_name,
        player_position
    ])

print(f"\n=== DRAFT (first 20 of {pick_count} picks) ===")
print(f"{'Rnd':>4} {'Pick':>4}  {'Team':<28} {'Player':<25} {'Pos'}")
print("-" * 72)
for r in draft_rows[:20]:
    print(f"{r[0]:>4} {r[1]:>4}  {r[2]:<28} {r[3]:<25} {r[4]}")
if len(draft_rows) > 20:
    print(f"  ... and {len(draft_rows) - 20} more picks")

# ── 3. WEEKLY MATCHUPS ───────────────────────────────────────────────────────
print("\nFetching weekly matchups...")
matchup_rows = []
for week in range(1, NUM_WEEKS + 1):
    print(f"  Week {week}...")
    try:
        data = yahoo_get(session, f"{BASE}/league/{LEAGUE_KEY}/scoreboard;week={week}?format=json")
        matchups = data["fantasy_content"]["league"][1]["scoreboard"]["0"]["matchups"]
        matchup_count = matchups["count"]
        for i in range(matchup_count):
            m = matchups[str(i)]["matchup"]
            teams = m["0"]["teams"]
            t1 = teams["0"]["team"]
            t2 = teams["1"]["team"]
            t1_name = next((x["name"] for x in t1[0] if isinstance(x, dict) and "name" in x), "")
            t2_name = next((x["name"] for x in t2[0] if isinstance(x, dict) and "name" in x), "")
            t1_pts = t1[1]["team_points"]["total"]
            t2_pts = t2[1]["team_points"]["total"]
            is_playoff = week >= PLAYOFF_START_WEEK
            matchup_rows.append([week, t1_name, t1_pts, t2_name, t2_pts, "Yes" if is_playoff else "No", ""])
    except Exception as e:
        print(f"  Week {week} error: {e}")

# ── PLAYOFF BRACKET LABELING ─────────────────────────────────────────────────
print("Labeling playoff rounds...")

# Get seeds from regular-season record (wins desc, points_for desc as tiebreaker)
seeded_teams = [row[0] for row in sorted(standings_rows, key=lambda r: (-int(r[2]), -float(r[5])))]
playoff_teams = set(seeded_teams[:6])
top_two = set(seeded_teams[:2])

# Track bracket progression
week14_winners = set()
week14_losers = set()
week15_winners = set()
week15_losers = set()

for i, row in enumerate(matchup_rows):
    week, t1, s1, t2, s2, is_playoff, _ = row
    if not is_playoff == "Yes":
        continue

    try:
        s1 = float(s1)
        s2 = float(s2)
    except:
        continue

    winner = t1 if s1 > s2 else t2
    loser = t1 if s1 < s2 else t2

    if week == PLAYOFF_START_WEEK:
        if t1 in top_two or t2 in top_two:
            label = "Consolation"
        elif t1 in playoff_teams and t2 in playoff_teams:
            label = "Quarterfinal"
            week14_winners.add(winner)
            week14_losers.add(loser)
        else:
            label = "Consolation"

    elif week == PLAYOFF_START_WEEK + 1:
        semifinal_teams = top_two | week14_winners
        if t1 in semifinal_teams and t2 in semifinal_teams:
            label = "Semifinal"
            week15_winners.add(winner)
            week15_losers.add(loser)
        else:
            label = "Consolation"

    elif week == PLAYOFF_START_WEEK + 2:
        if t1 in week15_winners and t2 in week15_winners:
            label = "Championship"
        elif t1 in week15_losers and t2 in week15_losers:
            label = "3rd Place"
        else:
            label = "Consolation"
    else:
        label = ""

    matchup_rows[i][6] = label

print("\n=== REGULAR SEASON MATCHUPS (Week 1 sample) ===")
print(f"{'Wk':>3}  {'Team 1':<28} {'Score 1':>8}  {'Team 2':<28} {'Score 2':>8}")
print("-" * 82)
for r in [r for r in matchup_rows if r[0] == 1]:
    print(f"{r[0]:>3}  {r[1]:<28} {float(r[2]):>8.2f}  {r[3]:<28} {float(r[4]):>8.2f}")

print("\n=== PLAYOFF MATCHUPS ===")
print(f"{'Wk':>3}  {'Team 1':<28} {'Score 1':>8}  {'Team 2':<28} {'Score 2':>8}  {'Round'}")
print("-" * 95)
for r in [r for r in matchup_rows if r[5] == "Yes"]:
    print(f"{r[0]:>3}  {r[1]:<28} {float(r[2]):>8.2f}  {r[3]:<28} {float(r[4]):>8.2f}  {r[6]}")

# ── 4. TRANSACTIONS ──────────────────────────────────────────────────────────
print("\nFetching transactions...")
transaction_rows = []

transaction_types = ["add", "drop", "trade"]

for t_type in transaction_types:
    print(f"  Fetching {t_type}s...")
    try:
        url = f"{BASE}/league/{LEAGUE_KEY}/transactions;type={t_type}?format=json"
        data = yahoo_get(session, url)
        transactions = data["fantasy_content"]["league"][1]["transactions"]
        t_count = transactions["count"]
        print(f"  Found {t_count} {t_type}s")
        for i in range(t_count):
            t = transactions[str(i)]["transaction"]
            t_info = t[0]
            t_players = t[1].get("players", {})
            p_count = t_players.get("count", 0)
            timestamp = t_info.get("timestamp", "")
            for j in range(p_count):
                p = t_players[str(j)]["player"]
                p_info = p[0]
                p_data = p[1].get("transaction_data", {})
                pname = ""
                for x in p_info:
                    if isinstance(x, dict) and "name" in x and isinstance(x["name"], dict):
                        pname = x["name"].get("full", "")
                if isinstance(p_data, list):
                    p_data = p_data[0] if p_data else {}
                source_team = p_data.get("source_team_name", "FA/Waiver")
                dest_team = p_data.get("destination_team_name", "Dropped")
                transaction_rows.append([
                    t_type.capitalize(),
                    timestamp,
                    pname,
                    source_team,
                    dest_team
                ])
        time.sleep(0.5)
    except Exception as e:
        print(f"  Error fetching {t_type}s: {e}")

print(f"\n=== TRANSACTIONS (first 20 of {len(transaction_rows)}) ===")
print(f"{'Type':<8} {'Timestamp':<12}  {'Player':<25} {'From':<25} {'To'}")
print("-" * 90)
for r in transaction_rows[:20]:
    print(f"{r[0]:<8} {str(r[1]):<12}  {r[2]:<25} {r[3]:<25} {r[4]}")
if len(transaction_rows) > 20:
    print(f"  ... and {len(transaction_rows) - 20} more transactions")

# ── 5. END OF SEASON ROSTERS ─────────────────────────────────────────────────
team_keys = list(team_key_to_name.keys())
print("\nFetching end of season rosters...")
roster_rows = []

for team_key in team_keys:
    team_name = team_key_to_name.get(team_key, team_key)
    try:
        url = f"{BASE}/team/{team_key}/roster;week={NUM_WEEKS}?format=json"
        data = yahoo_get(session, url)
        players = data["fantasy_content"]["team"][1]["roster"]["0"]["players"]
        p_count = players["count"]
        for i in range(p_count):
            p = players[str(i)]["player"][0]
            pname = ""
            pposition = ""
            pteam = ""
            pstatus = ""
            for x in p:
                if isinstance(x, dict) and "name" in x and isinstance(x["name"], dict):
                    pname = x["name"].get("full", "")
                if isinstance(x, dict) and "display_position" in x:
                    pposition = x["display_position"]
                if isinstance(x, dict) and "editorial_team_abbr" in x:
                    pteam = x["editorial_team_abbr"]
                if isinstance(x, dict) and "status" in x:
                    pstatus = x["status"]
            roster_rows.append([SEASON_YEAR, team_name, pname, pposition, pteam, pstatus])
        print(f"  {team_name}: {p_count} players")
        time.sleep(0.5)
    except Exception as e:
        print(f"  Error fetching roster for {team_name}: {e}")

print(f"\n=== END OF SEASON ROSTERS (first 20 of {len(roster_rows)}) ===")
print(f"{'Fantasy Team':<28} {'Player':<25} {'Pos':<6} {'NFL Team':<10} {'Status'}")
print("-" * 80)
for r in roster_rows[:20]:
    print(f"{r[1]:<28} {r[2]:<25} {r[3]:<6} {r[4]:<10} {r[5]}")
if len(roster_rows) > 20:
    print(f"  ... and {len(roster_rows) - 20} more roster spots")

print("\n=== LEAGUE SCORING SETTINGS ===")
print(f"{'Stat':<35} {'Pts':>6}")
print("-" * 43)
for r in settings_rows:
    print(f"{r[0]:<35} {r[1]:>6}")

# ── 6. WEEKLY LINEUPS + YAHOO FANTASY POINTS ─────────────────────────────────
# Fetch roster (for slot) and player points (from Yahoo) in two calls per team/week,
# joined by player_key.

LINEUP_HEADERS = ["Week", "Fantasy Team", "Slot", "Player", "Pos", "NFL Team", "Fantasy Pts"]

print("\nFetching weekly lineups + Yahoo fantasy points...")
player_week_rows = []

for week in range(1, NUM_WEEKS + 1):
    print(f"  Week {week}...")
    for team_key, team_name in team_key_to_name.items():
        try:
            # Call 1: week-specific roster → player_keys, slots, and player info
            roster_data = yahoo_get(session, f"{BASE}/team/{team_key}/roster;week={week}?format=json")
            roster_players = roster_data["fantasy_content"]["team"][1]["roster"]["0"]["players"]
            players_this_week = []  # (pk, slot, name, pos, nfl_team)
            for i in range(roster_players["count"]):
                p = roster_players[str(i)]["player"]
                info = p[0]
                p_selected = p[1].get("selected_position", [])
                pk    = next((x["player_key"] for x in info if isinstance(x, dict) and "player_key" in x), None)
                pname = next((x["name"]["full"] for x in info if isinstance(x, dict) and "name" in x and isinstance(x["name"], dict)), "")
                ppos  = next((x["display_position"] for x in info if isinstance(x, dict) and "display_position" in x), "")
                pteam = next((x["editorial_team_abbr"] for x in info if isinstance(x, dict) and "editorial_team_abbr" in x), "")
                slot = ""
                if isinstance(p_selected, list):
                    for s in p_selected:
                        if isinstance(s, dict) and "position" in s:
                            slot = s["position"]
                elif isinstance(p_selected, dict):
                    slot = p_selected.get("position", "")
                if pk:
                    players_this_week.append((pk, slot, pname, ppos, pteam))
            time.sleep(0.2)

            # Call 2: fetch points for exactly those player_keys
            keys_str = ",".join(pk for pk, *_ in players_this_week)
            pts_data = yahoo_get(session, f"{BASE}/league/{LEAGUE_KEY}/players;player_keys={keys_str}/stats;type=week;week={week}?format=json")
            pts_players = pts_data["fantasy_content"]["league"][1]["players"]
            pts_by_key = {}
            for i in range(pts_players["count"]):
                p = pts_players[str(i)]["player"]
                info = p[0]
                sb   = p[1] if len(p) > 1 else {}
                pk   = next((x["player_key"] for x in info if isinstance(x, dict) and "player_key" in x), "")
                pts_by_key[pk] = sb.get("player_points", {}).get("total", "")
            time.sleep(0.2)

            for pk, slot, pname, ppos, pteam in players_this_week:
                pts = pts_by_key.get(pk, "")
                player_week_rows.append([week, team_name, slot, pname, ppos, pteam, pts])

        except Exception as e:
            print(f"    Error {team_name} week {week}: {e}")

print(f"  {len(player_week_rows)} player-week records fetched.")

# Console preview: week 1, first team
preview_team = next((r[1] for r in player_week_rows if r[0] == 1), "")
print(f"\n=== WEEKLY LINEUPS + POINTS PREVIEW (Week 1 — {preview_team}) ===")
print(f"{'Slot':<6} {'Player':<25} {'Pos':<5} {'NFL':<5} {'Pts':>7}")
print("-" * 52)
for r in [x for x in player_week_rows if x[0] == 1 and x[1] == preview_team]:
    print(f"{r[2]:<6} {r[3]:<25} {r[4]:<5} {r[5]:<5} {str(r[6]):>7}")

# ── EXPORT TO EXCEL (pass --export to enable) ─────────────────────────────────
if EXPORT:
    print("\nWriting to Excel...")
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Standings"
    ws.append(["Team", "Manager", "Wins", "Losses", "Ties", "Points For", "Points Against", "Rank"])
    for row in standings_rows:
        ws.append(row)

    ws2 = wb.create_sheet("Draft")
    ws2.append(["Round", "Pick", "Team", "Player", "Position"])
    for row in draft_rows:
        ws2.append(row)

    ws3 = wb.create_sheet("Matchups")
    ws3.append(["Week", "Team 1", "Score 1", "Team 2", "Score 2", "Playoff", "Round"])
    for row in matchup_rows:
        ws3.append(row)

    ws5 = wb.create_sheet("Transactions")
    ws5.append(["Type", "Timestamp", "Player", "From", "To"])
    for row in transaction_rows:
        ws5.append(row)

    ws6 = wb.create_sheet("End of Season Rosters")
    ws6.append(["Season", "Fantasy Team", "Player", "Position", "NFL Team", "Status"])
    for row in roster_rows:
        ws6.append(row)

    ws7 = wb.create_sheet("Playoffs")
    ws7.append(["Week", "Team 1", "Score 1", "Team 2", "Score 2", "Round"])
    for row in matchup_rows:
        if row[0] >= PLAYOFF_START_WEEK:
            ws7.append([row[0], row[1], row[2], row[3], row[4], row[6]])

    ws_settings = wb.create_sheet("League Settings")
    ws_settings.append(["Stat", "Points"])
    for row in settings_rows:
        ws_settings.append(row)

    for week in range(1, NUM_WEEKS + 1):
        ws_week = wb.create_sheet(f"Week {week:02d}")
        ws_week.append(LINEUP_HEADERS)
        for row in [r for r in player_week_rows if r[0] == week]:
            ws_week.append(row)

    output = os.path.expanduser("~/ffhistory/ff_2011.xlsx")
    wb.save(output)
    print(f"Saved to {output}")
else:
    print("\nRun with --export to save to Excel.")
